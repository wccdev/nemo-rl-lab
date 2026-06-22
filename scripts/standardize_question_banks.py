#!/usr/bin/env python3
"""Standardize department question-bank Excel files for RAGFlow upload."""

from __future__ import annotations

import argparse
import json
import re
import shutil
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook

STANDARD_SHEETS = ("单选题", "多选题", "判断题", "填空题", "简答题")
CHOICE_SHEETS = {"单选题", "多选题"}
EXAMPLE_PATTERNS = (
    "示例导入前请删除",
    "（示例",
    "(示例",
    "这是单选题题目",
    "这是一道多选题",
    "这是一道判断题",
    "这是一道填空题",
    "这是一道简答题",
)

INSTRUCTIONS = {
    "单选题": (
        "单选题说明（说明部分请勿删除）：\n"
        "1.序号为必填项，请从1开始排序；\n"
        "2.单选题正确答案只有一个，请填写大写A B C D E F G（半角英文大写）；\n"
        "3.单选题最多支持12个选项。请勿删除和增加选项列（选项为空时无需填写）；\n"
        "4.请尽量避免特殊字符输入（表情、乱码），以免影响系统校验；\n"
        "5.系统会先校验模板中试题录入正确性，有错误时试题不会进行导入，修订后重新导入即可。\n"
        "6.每道题最多添加5个知识点，多个知识点用竖线“|”分割；\n"
        "7.其他题型（多选、判断、填空、简答）请在底部对应Sheet页录入；不需要录入其他题型时，请先删除其Sheet表中的示例行（表头保留）；\n"
        "8.试题内容不支持公式填充，请避免此操作，以免影响系统校验"
    ),
    "多选题": (
        "多选题说明（说明部分请勿删除）：\n"
        "1.序号为必填项，请从1开始排序；\n"
        "2.多选题正确答案至少一个，请填写大写A B C D E F G（半角英文大写）；\n"
        "3.多选题最多支持12个选项。请勿删除和增加选项列（选项为空时无需填写）；\n"
        "4.请尽量避免特殊字符输入（表情、乱码），以免影响系统校验；\n"
        "5.系统会先校验模板中试题录入正确性，有错误时试题不会进行导入，修订后重新导入即可。\n"
        "6.每道题最多添加5个知识点，多个知识点用竖线“|”分割；\n"
        "7.其他题型（单选、判断、填空、简答）请在底部对应Sheet页录入；不需要录入其他题型时，请先删除其Sheet表中的示例行（表头保留）；\n"
        "8.试题内容不支持公式填充，请避免此操作，以免影响系统校验"
    ),
    "判断题": (
        "判断题说明（说明部分请勿删除）：\n"
        "1.序号为必填项，请从1开始排序；\n"
        "2.判断题正确答案选择“对”或“错”；\n"
        "3.请尽量避免特殊字符输入（表情、乱码），以免影响系统校验；\n"
        "4.系统会先校验模板中试题录入正确性，有错误时试题不会进行导入，修订后重新导入即可。\n"
        "5.每道题最多添加5个知识点，多个知识点用竖线“|”分割；\n"
        "6.其他题型（单选、多选、填空、简答）请在底部对应Sheet页录入；不需要录入其他题型时，请先删除其Sheet表中的示例行（表头保留）；\n"
        "7.试题内容不支持公式填充，请避免此操作，以免影响系统校验"
    ),
    "填空题": (
        "填空题说明（说明部分请勿删除）：\n"
        "1.序号为必填项，请从1开始排序；\n"
        "2.填空题题目英文输入下划线“____”表示填空位置；\n"
        "3.请尽量避免特殊字符输入（表情、乱码），以免影响系统校验；\n"
        "4.系统会先校验模板中试题录入正确性，有错误时试题不会进行导入，修订后重新导入即可。\n"
        "5.每道题最多添加5个知识点，多个知识点用竖线“|”分割；\n"
        "6.其他题型（单选、多选、判断、简答）请在底部对应Sheet页录入；不需要录入其他题型时，请先删除其Sheet表中的示例行（表头保留）；\n"
        "7.试题内容不支持公式填充，请避免此操作，以免影响系统校验"
    ),
    "简答题": (
        "简答题说明（说明部分请勿删除）：\n"
        "1.序号为必填项，请从1开始排序；\n"
        "2.简答题可以设置多个关键词，多个关键词用竖线“|”分割；\n"
        "3.请尽量避免特殊字符输入（表情、乱码），以免影响系统校验；\n"
        "4.系统会先校验模板中试题录入正确性，有错误时试题不会进行导入，修订后重新导入即可。\n"
        "5.每道题最多添加5个知识点，多个知识点用竖线“|”分割；\n"
        "6.其他题型（单选、多选、判断、填空）请在底部对应Sheet页录入；不需要录入其他题型时，请先删除其Sheet表中的示例行（表头保留）；\n"
        "7.试题内容不支持公式填充，请避免此操作，以免影响系统校验"
    ),
}

CHOICE_HEADERS = (
    "序号（必填）",
    "题目(必填)",
    "正确答案(必填)",
    "选项",
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    "知识点",
    "难度*(必填)",
    "答案解析",
)
CHOICE_OPTION_LABELS = (
    "A",
    "B",
    "C",
    "D",
    "E(勿删)",
    "F(勿删)",
    "G(勿删)",
    "H(勿删)",
    "I(勿删)",
    "J(勿删)",
    "K(勿删)",
    "L(勿删)",
)
JUDGE_HEADERS = (
    "序号（必填）",
    "题目(必填)",
    "正确答案(必填)",
    "知识点",
    "难度*(必填)",
    "答案解析",
)
FILL_HEADERS = ("序号（必填）", "题目(必填)", "填空项")
FILL_BLANK_LABELS = tuple(f"空{i}" for i in range(1, 11))
SHORT_HEADERS = ("序号（必填）", "题目(必填)", "关键词")
SHORT_KEYWORD_LABELS = tuple(f"关键词{i}" for i in range(1, 11))


@dataclass
class Question:
    qtype: str
    question: str
    answer: str
    options: list[str] = field(default_factory=list)
    knowledge_points: str = ""
    difficulty: str = "简单"
    explanation: str = ""
    source_file: str = ""
    source_sheet: str = ""
    row_num: int = 0

    def dedupe_key(self) -> str:
        opts = "|".join(self.options)
        return f"{self.qtype}::{self.question.strip()}::{opts}::{self.answer.strip()}"


def norm_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("（必填）", "").replace("(必填)", "")
    text = text.replace("*", "").replace(" ", "")
    return text


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"[\u200b-\u200d\ufeff]", "", text)
    text = re.sub(r"\s+", " ", text)
    return text


def is_example_question(text: str) -> bool:
    return any(pattern in text for pattern in EXAMPLE_PATTERNS)


def source_title(filename: str) -> str:
    name = filename
    if name.startswith("20260611160325_"):
        name = name[len("20260611160325_") :]
    for suffix in (".xlsx", ".xls"):
        if name.endswith(suffix):
            name = name[: -len(suffix)]
    return name


def find_header_row(rows: list[tuple[Any, ...]]) -> int | None:
    for idx, row in enumerate(rows):
        cells = [norm_header(cell) for cell in row[:8]]
        if "序号" in cells and ("题目" in cells or "试题" in cells):
            return idx
    return None


def col_index(headers: list[str], *candidates: str) -> int | None:
    for candidate in candidates:
        for idx, header in enumerate(headers):
            if not header:
                continue
            if candidate in header or header in candidate:
                return idx
    return None


def normalize_choice_answer(raw: str, qtype: str) -> str | None:
    text = clean_text(raw).upper().replace(" ", "").replace(",", "").replace("，", "")
    if not text:
        return None
    letters = re.findall(r"[A-L]", text)
    if not letters:
        return None
    answer = "".join(dict.fromkeys(letters))
    if qtype == "单选题" and len(answer) != 1:
        return answer[:1]
    return answer


def normalize_judge_answer(raw: str) -> str | None:
    text = clean_text(raw).lower()
    if text in {"对", "正确", "true", "t", "yes", "y", "√", "是"}:
        return "对"
    if text in {"错", "错误", "false", "f", "no", "n", "×", "否", "不对"}:
        return "错"
    if text in {"a", "b"}:
        return "对" if text == "a" else "错"
    return None


def normalize_knowledge_points(raw: str) -> str:
    if not raw:
        return ""
    parts = [clean_text(part) for part in re.split(r"[|｜]", raw) if clean_text(part)]
    return "|".join(parts[:5])


def looks_like_question_bank(sheet_names: list[str]) -> bool:
    return bool(set(sheet_names) & set(STANDARD_SHEETS))


def parse_choice_sheet(
    ws,
    qtype: str,
    source_file: str,
) -> tuple[list[Question], list[str]]:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = find_header_row(rows)
    if header_idx is None:
        return [], ["missing_header"]

    headers = [norm_header(cell) for cell in rows[header_idx]]
    idx_q = col_index(headers, "题目")
    idx_a = col_index(headers, "正确答案", "答案")
    if idx_q is None or idx_a is None:
        return [], ["bad_header"]

    option_start = col_index(headers, "选项")
    if option_start is None:
        option_start = 3
    idx_kp = col_index(headers, "知识点")
    idx_diff = col_index(headers, "难度")
    idx_exp = col_index(headers, "答案解析", "解析")

    questions: list[Question] = []
    issues: list[str] = []

    for row_num, row in enumerate(rows[header_idx + 2 :], start=header_idx + 3):
        if not row:
            continue
        question = clean_text(row[idx_q] if idx_q < len(row) else None)
        if not question:
            continue
        if is_example_question(question):
            continue

        answer_raw = row[idx_a] if idx_a < len(row) else None
        answer = normalize_choice_answer(clean_text(answer_raw), qtype)
        if not answer:
            issues.append(f"{qtype}:row{row_num}:missing_answer")
            continue

        options: list[str] = []
        for offset in range(12):
            col = option_start + offset
            if col >= len(row):
                options.append("")
            else:
                options.append(clean_text(row[col]))

        valid_letters = {chr(ord("A") + i) for i, opt in enumerate(options) if opt}
        if valid_letters and any(letter not in valid_letters for letter in answer):
            issues.append(f"{qtype}:row{row_num}:answer_out_of_range")
            continue

        kp = normalize_knowledge_points(
            clean_text(row[idx_kp]) if idx_kp is not None and idx_kp < len(row) else ""
        )
        diff = clean_text(row[idx_diff]) if idx_diff is not None and idx_diff < len(row) else ""
        exp = clean_text(row[idx_exp]) if idx_exp is not None and idx_exp < len(row) else ""

        questions.append(
            Question(
                qtype=qtype,
                question=question,
                answer=answer,
                options=options,
                knowledge_points=kp,
                difficulty=diff or "简单",
                explanation=exp,
                source_file=source_file,
                source_sheet=qtype,
                row_num=row_num,
            )
        )
    return questions, issues


def parse_judge_sheet(ws, source_file: str) -> tuple[list[Question], list[str]]:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = find_header_row(rows)
    if header_idx is None:
        return [], ["missing_header"]

    headers = [norm_header(cell) for cell in rows[header_idx]]
    idx_q = col_index(headers, "题目")
    idx_a = col_index(headers, "正确答案", "答案")
    idx_kp = col_index(headers, "知识点")
    idx_diff = col_index(headers, "难度")
    idx_exp = col_index(headers, "答案解析", "解析")
    if idx_q is None or idx_a is None:
        return [], ["bad_header"]

    questions: list[Question] = []
    issues: list[str] = []

    for row_num, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if not row:
            continue
        question = clean_text(row[idx_q] if idx_q < len(row) else None)
        if not question:
            continue
        if is_example_question(question):
            continue
        if re.search(r"[A-L][\.、．)]", question):
            issues.append(f"判断题:row{row_num}:looks_like_mixed_choice")
            continue

        answer = normalize_judge_answer(row[idx_a] if idx_a < len(row) else "")
        if not answer:
            issues.append(f"判断题:row{row_num}:missing_answer")
            continue

        kp = normalize_knowledge_points(
            clean_text(row[idx_kp]) if idx_kp is not None and idx_kp < len(row) else ""
        )
        diff = clean_text(row[idx_diff]) if idx_diff is not None and idx_diff < len(row) else ""
        exp = clean_text(row[idx_exp]) if idx_exp is not None and idx_exp < len(row) else ""

        questions.append(
            Question(
                qtype="判断题",
                question=question,
                answer=answer,
                knowledge_points=kp,
                difficulty=diff or "简单",
                explanation=exp,
                source_file=source_file,
                source_sheet="判断题",
                row_num=row_num,
            )
        )
    return questions, issues


def parse_fill_sheet(ws, source_file: str) -> tuple[list[Question], list[str]]:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = find_header_row(rows)
    if header_idx is None:
        return [], ["missing_header"]

    headers = [norm_header(cell) for cell in rows[header_idx]]
    idx_q = col_index(headers, "题目")
    blank_start = col_index(headers, "填空项")
    idx_kp = col_index(headers, "知识点")
    idx_diff = col_index(headers, "难度")
    if idx_q is None:
        return [], ["bad_header"]
    if blank_start is None:
        blank_start = 2

    questions: list[Question] = []
    issues: list[str] = []

    for row_num, row in enumerate(rows[header_idx + 2 :], start=header_idx + 3):
        question = clean_text(row[idx_q] if idx_q < len(row) else None)
        if not question:
            continue
        if is_example_question(question):
            continue

        blanks = []
        for offset in range(10):
            col = blank_start + offset
            if col >= len(row):
                blanks.append("")
            else:
                blanks.append(clean_text(row[col]))
        answer = "|".join(part for part in blanks if part)
        if not answer:
            issues.append(f"填空题:row{row_num}:missing_answer")
            continue

        kp = normalize_knowledge_points(
            clean_text(row[idx_kp]) if idx_kp is not None and idx_kp < len(row) else ""
        )
        diff = clean_text(row[idx_diff]) if idx_diff is not None and idx_diff < len(row) else ""

        questions.append(
            Question(
                qtype="填空题",
                question=question,
                answer=answer,
                options=blanks,
                knowledge_points=kp,
                difficulty=diff or "简单",
                source_file=source_file,
                source_sheet="填空题",
                row_num=row_num,
            )
        )
    return questions, issues


def parse_short_sheet(ws, source_file: str) -> tuple[list[Question], list[str]]:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = find_header_row(rows)
    if header_idx is None:
        return [], ["missing_header"]

    headers = [norm_header(cell) for cell in rows[header_idx]]
    idx_q = col_index(headers, "题目")
    keyword_start = col_index(headers, "关键词")
    idx_kp = col_index(headers, "知识点")
    idx_diff = col_index(headers, "难度")
    if idx_q is None:
        return [], ["bad_header"]
    if keyword_start is None:
        keyword_start = 2

    questions: list[Question] = []
    issues: list[str] = []

    for row_num, row in enumerate(rows[header_idx + 2 :], start=header_idx + 3):
        question = clean_text(row[idx_q] if idx_q < len(row) else None)
        if not question:
            continue
        if is_example_question(question):
            continue

        keywords = []
        for offset in range(10):
            col = keyword_start + offset
            if col >= len(row):
                keywords.append("")
            else:
                keywords.append(clean_text(row[col]))
        answer = "|".join(part for part in keywords if part)
        if not answer:
            issues.append(f"简答题:row{row_num}:missing_keywords")
            continue

        kp = normalize_knowledge_points(
            clean_text(row[idx_kp]) if idx_kp is not None and idx_kp < len(row) else ""
        )
        diff = clean_text(row[idx_diff]) if idx_diff is not None and idx_diff < len(row) else ""

        questions.append(
            Question(
                qtype="简答题",
                question=question,
                answer=answer,
                options=keywords,
                knowledge_points=kp,
                difficulty=diff or "简单",
                source_file=source_file,
                source_sheet="简答题",
                row_num=row_num,
            )
        )
    return questions, issues


def parse_workbook(path: Path) -> tuple[list[Question], dict[str, Any]]:
    result: dict[str, Any] = {
        "file": path.name,
        "status": "ok",
        "issues": [],
        "questions": 0,
    }
    questions: list[Question] = []

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        result["status"] = "error"
        result["issues"].append(str(exc))
        return [], result

    if not looks_like_question_bank(wb.sheetnames):
        result["status"] = "not_question_bank"
        result["issues"].append(f"sheets={wb.sheetnames[:5]}")
        wb.close()
        return [], result

    parsers = {
        "单选题": lambda ws: parse_choice_sheet(ws, "单选题", path.name),
        "多选题": lambda ws: parse_choice_sheet(ws, "多选题", path.name),
        "判断题": lambda ws: parse_judge_sheet(ws, path.name),
        "填空题": lambda ws: parse_fill_sheet(ws, path.name),
        "简答题": lambda ws: parse_short_sheet(ws, path.name),
    }

    for sheet_name, parser in parsers.items():
        if sheet_name not in wb.sheetnames:
            continue
        sheet_questions, sheet_issues = parser(wb[sheet_name])
        questions.extend(sheet_questions)
        result["issues"].extend(sheet_issues)

    wb.close()
    result["questions"] = len(questions)
    if not questions:
        result["status"] = "empty_or_invalid"
    return questions, result


def write_choice_sheet(ws, qtype: str, questions: list[Question]) -> None:
    ws.cell(1, 1, INSTRUCTIONS[qtype])
    for col, value in enumerate(CHOICE_HEADERS, start=1):
        ws.cell(2, col, value)
    for col, value in enumerate(CHOICE_OPTION_LABELS, start=4):
        ws.cell(3, col, value)

    for idx, question in enumerate(questions, start=1):
        row = idx + 3
        ws.cell(row, 1, idx)
        ws.cell(row, 2, question.question)
        ws.cell(row, 3, question.answer)
        for opt_idx, option in enumerate(question.options[:12]):
            ws.cell(row, 4 + opt_idx, option or None)
        ws.cell(row, 16, question.knowledge_points or None)
        ws.cell(row, 17, question.difficulty or "简单")
        ws.cell(row, 18, question.explanation or None)


def write_judge_sheet(ws, questions: list[Question]) -> None:
    ws.cell(1, 1, INSTRUCTIONS["判断题"])
    for col, value in enumerate(JUDGE_HEADERS, start=1):
        ws.cell(2, col, value)

    for idx, question in enumerate(questions, start=1):
        row = idx + 2
        ws.cell(row, 1, idx)
        ws.cell(row, 2, question.question)
        ws.cell(row, 3, question.answer)
        ws.cell(row, 4, question.knowledge_points or None)
        ws.cell(row, 5, question.difficulty or "简单")
        ws.cell(row, 6, question.explanation or None)


def write_fill_sheet(ws, questions: list[Question]) -> None:
    ws.cell(1, 1, INSTRUCTIONS["填空题"])
    for col, value in enumerate(FILL_HEADERS, start=1):
        ws.cell(2, col, value)
    for col, value in enumerate(FILL_BLANK_LABELS, start=3):
        ws.cell(3, col, value)

    for idx, question in enumerate(questions, start=1):
        row = idx + 3
        ws.cell(row, 1, idx)
        ws.cell(row, 2, question.question)
        for blank_idx, blank in enumerate(question.options[:10]):
            ws.cell(row, 3 + blank_idx, blank or None)


def write_short_sheet(ws, questions: list[Question]) -> None:
    ws.cell(1, 1, INSTRUCTIONS["简答题"])
    for col, value in enumerate(SHORT_HEADERS, start=1):
        ws.cell(2, col, value)
    for col, value in enumerate(SHORT_KEYWORD_LABELS, start=3):
        ws.cell(3, col, value)

    for idx, question in enumerate(questions, start=1):
        row = idx + 3
        ws.cell(row, 1, idx)
        ws.cell(row, 2, question.question)
        for kw_idx, keyword in enumerate(question.options[:10]):
            ws.cell(row, 3 + kw_idx, keyword or None)


def write_merged_workbook(path: Path, grouped: dict[str, list[Question]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    writers = {
        "单选题": lambda ws, items: write_choice_sheet(ws, "单选题", items),
        "多选题": lambda ws, items: write_choice_sheet(ws, "多选题", items),
        "判断题": write_judge_sheet,
        "填空题": write_fill_sheet,
        "简答题": write_short_sheet,
    }

    for sheet_name in STANDARD_SHEETS:
        items = grouped.get(sheet_name, [])
        ws = wb.create_sheet(sheet_name)
        writers[sheet_name](ws, items)

    wb.save(path)


def format_question_markdown(question: Question) -> str:
    lines = [
        f"## [{question.qtype}] {question.question}",
        "",
        f"- 来源文件: {source_title(question.source_file)}",
        f"- 来源Sheet: {question.source_sheet}",
        f"- 难度: {question.difficulty}",
    ]
    if question.knowledge_points:
        lines.append(f"- 知识点: {question.knowledge_points}")

    if question.qtype in CHOICE_SHEETS:
        lines.append("")
        lines.append("### 选项")
        for idx, option in enumerate(question.options):
            if not option:
                continue
            letter = chr(ord("A") + idx)
            lines.append(f"- {letter}. {option}")
    elif question.qtype == "填空题":
        lines.append("")
        lines.append("### 填空答案")
        for idx, blank in enumerate(question.options, start=1):
            if blank:
                lines.append(f"- 空{idx}: {blank}")
    elif question.qtype == "简答题":
        lines.append("")
        lines.append("### 参考答案关键词")
        for idx, keyword in enumerate(question.options, start=1):
            if keyword:
                lines.append(f"- 关键词{idx}: {keyword}")

    lines.extend(["", f"**正确答案**: {question.answer}", ""])
    if question.explanation:
        lines.extend([f"**解析**: {question.explanation}", ""])
    return "\n".join(lines)


def write_ragflow_outputs(output_dir: Path, grouped: dict[str, list[Question]]) -> None:
    ragflow_dir = output_dir / "ragflow"
    ragflow_dir.mkdir(parents=True, exist_ok=True)

    all_questions = [q for sheet in STANDARD_SHEETS for q in grouped.get(sheet, [])]
    by_source: dict[str, list[Question]] = defaultdict(list)
    for question in all_questions:
        by_source[source_title(question.source_file)].append(question)

    combined_lines = [
        "# 格科 Fab 培训题库知识库",
        "",
        f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"题目总数: {len(all_questions)}",
        "",
        "本文件用于 RAGFlow 知识库上传。每道题以二级标题分隔，便于检索召回。",
        "",
    ]
    for question in all_questions:
        combined_lines.append(format_question_markdown(question))
    (ragflow_dir / "all_questions.md").write_text("\n".join(combined_lines), encoding="utf-8")

    per_source_dir = ragflow_dir / "by_source"
    per_source_dir.mkdir(exist_ok=True)
    for source, questions in sorted(by_source.items()):
        safe_name = re.sub(r'[\\/:*?"<>|]', "_", source)[:120]
        lines = [f"# {source}", "", f"题目数量: {len(questions)}", ""]
        for question in questions:
            lines.append(format_question_markdown(question))
        (per_source_dir / f"{safe_name}.md").write_text("\n".join(lines), encoding="utf-8")

    jsonl_path = ragflow_dir / "questions.jsonl"
    with jsonl_path.open("w", encoding="utf-8") as handle:
        for question in all_questions:
            payload = asdict(question)
            payload["source_title"] = source_title(question.source_file)
            payload["rag_text"] = format_question_markdown(question)
            handle.write(json.dumps(payload, ensure_ascii=False) + "\n")

    csv_path = ragflow_dir / "questions_flat.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        import csv

        writer = csv.writer(handle)
        writer.writerow(
            [
                "序号",
                "题型",
                "题目",
                "正确答案",
                "选项A",
                "选项B",
                "选项C",
                "选项D",
                "选项E",
                "选项F",
                "选项G",
                "选项H",
                "选项I",
                "选项J",
                "选项K",
                "选项L",
                "知识点",
                "难度",
                "来源文件",
                "来源Sheet",
            ]
        )
        for idx, question in enumerate(all_questions, start=1):
            options = question.options + [""] * (12 - len(question.options))
            writer.writerow(
                [
                    idx,
                    question.qtype,
                    question.question,
                    question.answer,
                    *options[:12],
                    question.knowledge_points,
                    question.difficulty,
                    source_title(question.source_file),
                    question.source_sheet,
                ]
            )


def write_report(output_dir: Path, file_reports: list[dict[str, Any]], grouped: dict[str, list[Question]], rejected_counter: Counter) -> None:
    included_files = [item for item in file_reports if item["questions"] > 0]
    excluded_files = [item for item in file_reports if item["questions"] == 0]

    lines = [
        "# 题库标准化处理报告",
        "",
        f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## 汇总",
        "",
        f"- 扫描文件数: {len(file_reports)}",
        f"- 有效题库文件: {len(included_files)}",
        f"- 排除文件: {len(excluded_files)}",
        "",
        "## 题型统计",
        "",
    ]
    for sheet_name in STANDARD_SHEETS:
        lines.append(f"- {sheet_name}: {len(grouped.get(sheet_name, []))}")

    lines.extend(["", "## 剔除原因统计", ""])
    for reason, count in rejected_counter.most_common():
        lines.append(f"- {reason}: {count}")

    lines.extend(["", "## 已纳入题库文件", ""])
    for item in sorted(included_files, key=lambda x: -x["questions"]):
        lines.append(f"- {source_title(item['file'])}: {item['questions']} 题")

    lines.extend(["", "## 排除文件", ""])
    for item in sorted(excluded_files, key=lambda x: x["status"]):
        issue_preview = "; ".join(item["issues"][:2]) if item["issues"] else ""
        lines.append(f"- [{item['status']}] {item['file']} {issue_preview}")

    (output_dir / "processing_report.md").write_text("\n".join(lines), encoding="utf-8")
    (output_dir / "processing_report.json").write_text(
        json.dumps(
            {
                "summary": {
                    "files_scanned": len(file_reports),
                    "files_included": len(included_files),
                    "files_excluded": len(excluded_files),
                    "questions_by_type": {k: len(v) for k, v in grouped.items()},
                    "rejected_reasons": dict(rejected_counter),
                },
                "files": file_reports,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def dedupe_questions(questions: list[Question]) -> tuple[list[Question], int]:
    seen: set[str] = set()
    unique: list[Question] = []
    duplicates = 0
    for question in questions:
        key = question.dedupe_key()
        if key in seen:
            duplicates += 1
            continue
        seen.add(key)
        unique.append(question)
    return unique, duplicates


def main() -> None:
    parser = argparse.ArgumentParser(description="Standardize question bank Excel files.")
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path.home() / "Downloads/xls",
        help="Directory containing source Excel files",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path.home() / "Downloads/xls-ragflow-ready",
        help="Directory for standardized outputs",
    )
    parser.add_argument(
        "--keep-duplicates",
        action="store_true",
        help="Keep duplicate questions instead of deduplicating",
    )
    args = parser.parse_args()

    input_dir: Path = args.input_dir
    output_dir: Path = args.output_dir
    if output_dir.exists():
        shutil.rmtree(output_dir)
    output_dir.mkdir(parents=True)

    files = sorted(
        path
        for path in input_dir.glob("*.xlsx")
        if not path.name.startswith("~$") and not path.name.startswith(".~")
    )

    all_questions: list[Question] = []
    file_reports: list[dict[str, Any]] = []
    rejected_counter: Counter = Counter()

    for path in files:
        questions, report = parse_workbook(path)
        for issue in report["issues"]:
            if issue in {"missing_header", "bad_header"}:
                rejected_counter[issue] += 1
            elif ":" in issue:
                rejected_counter[issue.split(":", 1)[1]] += 1
            else:
                rejected_counter[issue] += 1
        all_questions.extend(questions)
        file_reports.append(report)

    duplicate_count = 0
    if not args.keep_duplicates:
        all_questions, duplicate_count = dedupe_questions(all_questions)
        rejected_counter["duplicate"] = duplicate_count

    grouped: dict[str, list[Question]] = {sheet: [] for sheet in STANDARD_SHEETS}
    for question in all_questions:
        grouped[question.qtype].append(question)

    write_merged_workbook(output_dir / "merged_question_bank.xlsx", grouped)
    write_ragflow_outputs(output_dir, grouped)
    write_report(output_dir, file_reports, grouped, rejected_counter)

    print(json.dumps(
        {
            "output_dir": str(output_dir),
            "files_scanned": len(files),
            "questions_total": len(all_questions),
            "questions_by_type": {k: len(v) for k, v in grouped.items()},
            "duplicates_removed": duplicate_count,
        },
        ensure_ascii=False,
        indent=2,
    ))


if __name__ == "__main__":
    main()
